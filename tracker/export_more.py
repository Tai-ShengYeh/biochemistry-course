#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""產出 Gimkit CSV 與 Wordwall 貼上用對照表。

平台差異(查證後):
  Kahoot / Wayground —— 匯入時指定正解「位置」→ 需要均勻分布(由 balance_options.py 處理)
  Gimkit —— CSV 格式為「正解欄 + 誤答欄」,平台自己隨機排列 → 位置分布不適用
  Wordwall —— 沒有匯入功能(官方明示),只能手動貼上 → 產出對照表供逐欄複製
              且 Wordwall 的強項是 Match up 配對題,不是選擇題
"""
import csv, glob, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BANKS='../banks'; GLOSS='../glossary_ids.csv'
OUT_G=os.path.join(BANKS,'gimkit'); OUT_W=os.path.join(BANKS,'wordwall')
os.makedirs(OUT_G,exist_ok=True); os.makedirs(OUT_W,exist_ok=True)
F="Arial"

TOPIC={2:"水、pH 與緩衝溶液",3:"胺基酸與胜肽",4:"蛋白質三維結構",5:"蛋白質功能",
 6:"酵素(一)",7:"酵素(二)",8:"碳水化合物",10:"脂質與生物膜",11:"核苷酸與核酸",
 12:"生物能量學",13:"糖解作用",14:"檸檬酸循環",15:"氧化磷酸化",16:"脂肪酸與胺基酸代謝",
 17:"代謝整合與激素調節",1:"導論:細胞與生物分子"}

# ---------- Gimkit ----------
ng=0
for f in sorted(glob.glob(f'{BANKS}/W*_題庫_主檔.csv')):
    wk=os.path.basename(f)[1:3]
    rows=list(csv.DictReader(open(f,encoding='utf-8-sig')))
    p=os.path.join(OUT_G,f'W{wk}_Gimkit.csv')
    with open(p,'w',newline='',encoding='utf-8-sig') as fh:
        w=csv.writer(fh)
        w.writerow(["Question","Correct Answer","Incorrect Answer 1","Incorrect Answer 2","Incorrect Answer 3"])
        for r in rows:
            opts=[r[f'opt{i}'] for i in range(1,5)]
            ci=int(r['correct'])-1
            correct=opts[ci]; wrong=[o for i,o in enumerate(opts) if i!=ci]
            w.writerow([r['question'],correct]+wrong)
    ng+=len(rows)
print(f"Gimkit:{len(glob.glob(f'{OUT_G}/*.csv'))} 週,{ng} 題(正解一律放 Correct Answer 欄)")

# ---------- Wordwall(Match up 配對:中文 ↔ 越南語)----------
gl=list(csv.DictReader(open(GLOSS,encoding='utf-8-sig')))
byw={}
for r in gl: byw.setdefault(int(r['week']),[]).append(r)

wb=Workbook(); wb.remove(wb.active)
thin=Side(style="thin",color="BFD7DE"); BD=Border(left=thin,right=thin,top=thin,bottom=thin)
nw=0
for wk in sorted(byw):
    ws=wb.create_sheet(f"W{wk:02d}")
    ws['A1']=f"第 {wk} 週 · {TOPIC.get(wk,'')} — Wordwall Match up 貼上用"
    ws['A1'].font=Font(name=F,size=13,bold=True,color="1F4E5F")
    ws['A2']="Wordwall 沒有匯入功能。建立 Match up 活動後,把下面兩欄逐格複製貼上。"
    ws['A2'].font=Font(name=F,size=10,italic=True,color="C8532B")
    for i,h in enumerate(["Keyword(越南語)","Definition(中文)"],1):
        c=ws.cell(4,i,h); c.font=Font(name=F,size=11,bold=True,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor="1F4E5F"); c.alignment=Alignment(horizontal="center"); c.border=BD
    for j,r in enumerate(byw[wk],5):
        ws.cell(j,1,r['vi']).border=BD; ws.cell(j,2,r['zh']).border=BD
        for c in (1,2): ws.cell(j,c).font=Font(name=F,size=11)
        nw+=1
    ws.column_dimensions['A'].width=32; ws.column_dimensions['B'].width=22
    ws.sheet_view.showGridLines=False
wb.save(os.path.join(OUT_W,'Wordwall_Matchup_貼上用.xlsx'))
# 另存純 TSV,方便整欄複製
for wk in sorted(byw):
    with open(os.path.join(OUT_W,f'W{wk:02d}_Wordwall_matchup.tsv'),'w',encoding='utf-8') as fh:
        for r in byw[wk]: fh.write(f"{r['vi']}\t{r['zh']}\n")
print(f"Wordwall:{len(byw)} 週,{nw} 組配對(xlsx 分頁 + 每週 TSV)")
