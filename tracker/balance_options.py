#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
balance_options.py — 讓選擇題的正解位置平均分布,並重新產出各平台匯入檔。

為什麼需要:出題時人會不自覺把正解放在前面。第1週原始版本裡,
選項4 一次都沒對過 —— 學生學會「不選第四個」就白賺分數。

用法:
  python balance_options.py ../W01_題庫_主檔.csv
  python balance_options.py ../W02_題庫_主檔.csv --seed 42

會做三件事:
  1. 重排每題的選項順序,讓正解位置盡量平均(且不出現連續3題同位置)
  2. 覆寫題庫主檔
  3. 重新產出 Kahoot .xlsx 與 Wayground .csv
"""
import csv, sys, random, collections, argparse, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

KAHOOT_HDR = ["Question - max 95 characters","Answer 1 - max 60 characters","Answer 2 - max 60 characters",
              "Answer 3 - max 60 characters","Answer 4 - max 60 characters",
              "Time limit (sec) – 5, 10, 20, 30, 60, 120","Correct answer(s) - choose at least one"]
TIME_LIMIT = 30      # A2 學生:不要用預設的 20 秒

def make_targets(n, k=4, seed=0, max_run=2):
    """產生 n 個正解位置,各位置次數盡量相等,且不出現連續 max_run+1 題同位置。
    不用 1,2,3,4,1,2,3,4 這種循環 —— 那雖然平均,但學生會看出規律。"""
    base = [(i % k) + 1 for i in range(n)]
    rng = random.Random(seed)
    for _ in range(10000):
        rng.shuffle(base)
        ok = True
        for i in range(len(base) - max_run):
            if len(set(base[i:i + max_run + 1])) == 1:
                ok = False; break
        if ok:
            return base
    return base   # 極端情況(n 很小)退回未檢查版本

def rebalance(rows, seed=0):
    targets = make_targets(len(rows), 4, seed)
    rng = random.Random(seed + 1)
    for r, tgt in zip(rows, targets):
        opts = [r["opt1"], r["opt2"], r["opt3"], r["opt4"]]
        correct_text = opts[int(r["correct"]) - 1]
        others = [o for i, o in enumerate(opts) if i != int(r["correct"]) - 1]
        rng.shuffle(others)                     # 誘答也重排,避免固定搭配
        new = others[:]
        new.insert(tgt - 1, correct_text)
        for i in range(4):
            r[f"opt{i+1}"] = new[i]
        r["correct"] = str(tgt)
        assert new[tgt - 1] == correct_text
    return rows

def report(rows, label):
    c = collections.Counter(r["correct"] for r in rows)
    print(f"\n{label}")
    for k in "1234":
        n = c.get(k, 0)
        print(f"  選項{k}: {n:2d} 次  {'■' * n}")
    print("  順序:", " ".join(r["correct"] for r in rows))
    return c

def write_kahoot(rows, path):
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    for i, h in enumerate(KAHOOT_HDR, 1):
        c = ws.cell(1, i, h); c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="46178F"); c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 45
    bad = []
    for r, row in enumerate(rows, 2):
        if len(row["question"]) > 95: bad.append(("題幹", row["term_id"], len(row["question"])))
        for i in range(1, 5):
            v = row[f"opt{i}"]
            if len(v) > 60: bad.append(("選項", row["term_id"], len(v)))
            ws.cell(r, 1 + i, v)
        ws.cell(r, 1, row["question"]); ws.cell(r, 6, TIME_LIMIT); ws.cell(r, 7, row["correct"])
    for col, w in zip("ABCDEFG", [46, 22, 22, 22, 22, 14, 14]): ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row: c.font = Font(name="Arial", size=11); c.alignment = Alignment(vertical="center")
    wb.save(path)
    return bad

def write_wayground(rows, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Question Text","Question Type","Option 1","Option 2","Option 3","Option 4",
                    "Correct Answer","Time in seconds","Tag"])
        for r in rows:
            w.writerow([r["question"],"Multiple Choice",r["opt1"],r["opt2"],r["opt3"],r["opt4"],
                        r["correct"],TIME_LIMIT,r["term_id"]])

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("bank"); ap.add_argument("--seed", type=int, default=0)
    a = ap.parse_args()
    rows = list(csv.DictReader(open(a.bank, encoding="utf-8-sig")))
    fields = list(rows[0].keys())
    report(rows, "修正前")
    rows = rebalance(rows, a.seed)
    c = report(rows, "修正後")
    assert max(c.values()) - min(c.values()) <= 1, "分布仍不平均"
    assert len(c) == 4, "有選項位置從未當過正解"

    with open(a.bank, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields); w.writeheader(); w.writerows(rows)
    d = os.path.dirname(os.path.abspath(a.bank))
    tag = os.path.basename(a.bank).split("_")[0]
    kp = os.path.join(d, f"{tag}_Kahoot匯入.xlsx"); wp = os.path.join(d, f"{tag}_Wayground_貼上用.csv")
    bad = write_kahoot(rows, kp); write_wayground(rows, wp)
    print(f"\n字元數檢查:{'全部通過' if not bad else bad}")
    print(f"已更新:\n  {a.bank}\n  {kp}\n  {wp}")

if __name__ == "__main__":
    main()
